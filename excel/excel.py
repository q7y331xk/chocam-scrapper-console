from openpyxl import Workbook
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re
from config import EXCEL_SAVE_PATH

def write_excel(columns, rows, title="default"):
    row_length = len(rows)
    col_length = len(columns)
    print(f"rows: {row_length} / cols: {col_length}")
    wb = Workbook()
    ws = wb.active
    ws.title = title
    k = 0
    while (k < col_length):
        ws.cell(row = 1, column = k + 1, value = columns[k])
        k = k + 1
    i = 0
    while (i < row_length):
        row = rows[i]
        j = 0
        while (j < col_length):
            item = row[j]
            # ws.cell(row = i + 1 + 1, column = j + 1).value = item
            item_str = str(item)
            ws.cell(row = i + 1 + 1, column = j + 1).value = ILLEGAL_CHARACTERS_RE.sub(r'', item_str)
            j = j + 1
        i = i + 1
    wb.save(EXCEL_SAVE_PATH)
    print(f'write excel in "{EXCEL_SAVE_PATH}"')


def write_exist_excel(sellings):
    col_length = len(sellings[0])
    wb = openpyxl.load_workbook(EXCEL_SAVE_PATH)
    ws = wb.active
    i = 150001
    while (i < 200000):
        row = sellings[i]
        j = 0
        while (j < col_length):
            item = row[j]
            item_str = str(item)
            ws.cell(row = i + 1 + 1, column = j + 1).value = ILLEGAL_CHARACTERS_RE.sub(r'', item_str)
            j = j + 1
        i = i + 1
    wb.save(EXCEL_SAVE_PATH)
    print(f'write excel in "{EXCEL_SAVE_PATH}"')